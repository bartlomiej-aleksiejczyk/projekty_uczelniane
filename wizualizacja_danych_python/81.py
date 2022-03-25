import pandas as pd
import numpy as np
import xlrd
import openpyxl

xlsx = pd.ExcelFile('imiona.xlsx')
wb = openpyxl.load_workbook('imiona.xlsx')
ws = wb['Arkusz1']
#print(ws['A1'].value)
#print(ws['A1'].value)
arkusz=pd.read_excel('imiona.xlsx', sheet_name='Arkusz1')
#df.to_excel('imiona.xlsx',  sheet_name='Arkusz1')
#for x in range(1,14619):
#    ws.cell(row=x, column=y)
#ws.cell(row=x, column=y)
print((arkusz.dtypes))
print (arkusz.query('Liczba >1000 '))
print (arkusz.query('Imie =="BARTŁOMIEJ" '))
print (arkusz.query('Imie =="BARTŁOMIEJ" '))
grupowany_po_roku=((arkusz.query('2000<=Rok<=2006')).groupby(["Rok"])["Liczba"].sum())
grupowany_po_urodzeniu=((arkusz.groupby(["Plec"])["Liczba"].sum()))
print(grupowany_po_roku)
grupowany_po_roku =arkusz.groupby("Liczba")['Rok'].sum()
print(grupowany_po_roku)
print(grupowany_po_urodzeniu)
najpopularniejsze_w_roku=(arkusz.groupby(["Plec","Rok","Imie"], as_index=False, sort=False)["Liczba"].sum())
najpopularniejsze_w_roku=(najpopularniejsze_w_roku.groupby(["Plec","Rok"], as_index=False, sort=False)["Liczba"].max())
najpopularniejsze_w_roku = najpopularniejsze_w_roku.reset_index()
print(najpopularniejsze_w_roku)
#a = grupowany_po_roku[grupowany_po_roku > 1999]
#print(a)
#print(arkusz["Liczba"].sum())
#grouped = arkusz.groupby(['Kontynent'])
#print(grouped.get_group('Europa'))
#print(df.isin(szukaj))
#print(df[((df.Populacja > 1000000) & (df.index.isin([0,2])))])
#rok = pd.read_excel(xlsx,'Rok')
#imie = pd.read_excel(xlsx,'Imie')
#plec = pd.read_excel(xlsx,'Plec')
#print(rok)